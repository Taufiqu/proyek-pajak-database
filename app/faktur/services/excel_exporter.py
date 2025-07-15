import os
import tempfile
import requests
from flask import send_file, jsonify
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, Side

SUPABASE_URL = os.getenv("SUPABASE_URL")
SUPABASE_KEY = os.getenv("SUPABASE_SERVICE_ROLE_KEY")
HEADERS = {
    "apikey": SUPABASE_KEY,
    "Authorization": f"Bearer {SUPABASE_KEY}",
}

def generate_excel_export(_):
    try:
        # === 1. Fetch Data dari Supabase ===
        def fetch_data(table, jenis_label):
            url = f"{SUPABASE_URL}/rest/v1/{table}?select=tanggal,keterangan,npwp_lawan_transaksi,nama_lawan_transaksi,no_faktur,dpp,ppn"
            response = requests.get(url, headers=HEADERS)
            if response.status_code != 200:
                raise ValueError(f"Gagal fetch data {jenis_label}: {response.text}")
            return [
                {
                    "tanggal": row["tanggal"],
                    "jenis": jenis_label,
                    "keterangan": row["keterangan"],
                    "npwp": row["npwp_lawan_transaksi"],
                    "nama": row["nama_lawan_transaksi"],
                    "faktur": row["no_faktur"],
                    "dpp": float(row["dpp"]),
                    "ppn": float(row["ppn"]),
                }
                for row in response.json()
            ]

        masukan = fetch_data("ppn_masukan", "PPN MASUKAN")
        keluaran = fetch_data("ppn_keluaran", "PPN KELUARAN")
        data_rows = masukan + keluaran

        # === 2. Siapkan Workbook dari Template ===
        BASE_DIR = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", ".."))
        template_path = os.path.normpath(os.path.join(BASE_DIR, "templates", "rekap_template.xlsx"))
        wb = load_workbook(template_path)
        ws = wb.active

        # === 3. Tulis Data ke Sheet ===
        headers = [
            "Tanggal", "Jenis", "Keterangan", "NPWP Rekanan",
            "Nama Rekanan", "No Faktur", "DPP", "PPN", "Jumlah"
        ]
        for col, title in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=title)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")

        thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin")
        )

        for i, item in enumerate(data_rows, start=2):
            values = [
                item["tanggal"],
                item["jenis"],
                item["keterangan"],
                item["npwp"],
                item["nama"],
                item["faktur"],
                item["dpp"],
                item["ppn"],
                item["dpp"] + item["ppn"]
            ]
            for j, value in enumerate(values, start=1):
                cell = ws.cell(row=i, column=j, value=value)
                cell.alignment = Alignment(vertical="center")
                cell.border = thin_border
                if isinstance(value, float) and j in [7, 8, 9]:
                    cell.number_format = "#,##0.00"

        # === 4. Simpan dan Return ===
        temp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
        wb.save(temp.name)
        return send_file(temp.name, as_attachment=True, download_name="rekap_pajak.xlsx")

    except Exception as e:
        print(f"[❌ ERROR /api/export] {e}")
        return jsonify(error=str(e)), 500
