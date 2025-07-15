import os
import tempfile
import requests
from flask import send_file, jsonify
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, Side

# Ambil environment Supabase
SUPABASE_URL = os.getenv("SUPABASE_URL")
SUPABASE_KEY = os.getenv("SUPABASE_SERVICE_ROLE_KEY")

HEADERS = {
    "apikey": SUPABASE_KEY,
    "Authorization": f"Bearer {SUPABASE_KEY}",
    "Content-Type": "application/json"
}

def generate_excel_bukti_setor_export():
    try:
        # 📥 Tarik data bukti setor dari Supabase
        response = requests.get(
            f"{SUPABASE_URL}/rest/v1/bukti_setor?select=*&order=tanggal.desc",
            headers=HEADERS
        )

        if response.status_code != 200:
            return jsonify(error="Gagal mengambil data dari Supabase"), 500

        data = response.json()
        print(f"[DEBUG] Jumlah data bukti setor: {len(data)}")

        # 📁 Cek dan buka template
        BASE_DIR = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", ".."))
        template_path = os.path.join(BASE_DIR, "templates", "rekap_template_bukti_setor.xlsx")
        if not os.path.exists(template_path):
            return jsonify(error=f"Template tidak ditemukan di path: {template_path}"), 404

        wb = load_workbook(template_path)
        ws = wb.active

        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # 📝 Tulis data ke Excel
        start_row = 2
        for idx, item in enumerate(data, start=start_row):
            ws.cell(row=idx, column=1, value=item["tanggal"]).border = thin_border
            ws.cell(row=idx, column=2, value=item["kode_setor"]).border = thin_border

            jumlah_cell = ws.cell(row=idx, column=3, value=float(item["jumlah"]))
            jumlah_cell.number_format = "#,##0.00"
            jumlah_cell.border = thin_border

            ws.cell(row=idx, column=4, value=item["created_at"]).border = thin_border

        # 💾 Simpan ke file sementara
        temp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
        wb.save(temp.name)
        print(f"[✅] Export Bukti Setor berhasil ke: {temp.name}")

        return send_file(temp.name, as_attachment=True, download_name="rekap_bukti_setor.xlsx")

    except Exception as e:
        print(f"[❌] Error generate Excel Bukti Setor: {e}")
        return jsonify({"error": str(e)}), 500
